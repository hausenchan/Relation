#!/usr/bin/env python3

import argparse
import json
from datetime import date, datetime, time
from decimal import Decimal

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def parse_args():
    parser = argparse.ArgumentParser(description="Extract workbook rows from .xlsx into JSON.")
    parser.add_argument("--input", required=True, help="Input .xlsx path")
    parser.add_argument("--output", required=True, help="Output JSON path")
    return parser.parse_args()


def normalize_value(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def is_non_empty(value):
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def header_score(row_values):
    non_empty_values = [value for value in row_values if is_non_empty(value)]
    if len(non_empty_values) < 2:
        return -1
    string_like = sum(1 for value in non_empty_values if isinstance(value, str))
    return len(non_empty_values) * 10 + string_like


def detect_header_row(rows, scan_limit=12):
    best_index = 0
    best_score = -1
    for index, row_values in enumerate(rows[:scan_limit]):
        score = header_score(row_values)
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def ensure_headers(raw_headers):
    headers = []
    used = {}
    for index, value in enumerate(raw_headers, start=1):
        header = str(value).strip() if is_non_empty(value) else f"column_{index}"
        base = header
        suffix = 2
        while header in used:
            header = f"{base}_{suffix}"
            suffix += 1
        used[header] = True
        headers.append(header)
    return headers


def build_row_object(headers, row_values):
    record = {}
    for header, value in zip(headers, row_values):
        record[header] = value
    return record


def extract_sheet(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([normalize_value(value) for value in row])

    if not rows:
        return {
            "sheet_name": ws.title,
            "title": None,
            "title_row_index": None,
            "header_row_index": None,
            "first_data_row_index": None,
            "row_count": 0,
            "column_count": 0,
            "non_empty_row_count": 0,
            "headers": [],
            "header_map": [],
            "merged_ranges": [],
            "rows": [],
        }

    header_row_zero = detect_header_row(rows)
    header_row_index = header_row_zero + 1
    raw_headers = rows[header_row_zero]
    headers = ensure_headers(raw_headers)

    title_row_index = None
    title = None
    if header_row_zero > 0:
        title_values = [value for value in rows[header_row_zero - 1] if is_non_empty(value)]
        if len(title_values) == 1:
            title_row_index = header_row_zero
            title = title_values[0]

    data_rows = []
    non_empty_row_count = 0
    for offset, row_values in enumerate(rows[header_row_zero + 1 :], start=header_row_index + 1):
        if not any(is_non_empty(value) for value in row_values):
            continue
        non_empty_row_count += 1
        data_rows.append(
            {
                "row_index": offset,
                "values": row_values,
                "record": build_row_object(headers, row_values),
            }
        )

    merged_ranges = [str(cell_range) for cell_range in ws.merged_cells.ranges]

    return {
        "sheet_name": ws.title,
        "title": title,
        "title_row_index": title_row_index,
        "header_row_index": header_row_index,
        "first_data_row_index": header_row_index + 1,
        "row_count": len(rows),
        "column_count": max((len(row) for row in rows), default=0),
        "non_empty_row_count": non_empty_row_count,
        "headers": headers,
        "header_map": [
            {
                "column_index": index,
                "column_letter": get_column_letter(index),
                "header": header,
            }
            for index, header in enumerate(headers, start=1)
        ],
        "merged_ranges": merged_ranges,
        "rows": data_rows,
    }


def main():
    args = parse_args()
    workbook = load_workbook(args.input, data_only=True)
    result = {
        "workbook": {
            "sheet_names": workbook.sheetnames,
            "sheet_count": len(workbook.sheetnames),
        },
        "sheets": [extract_sheet(workbook[sheet_name]) for sheet_name in workbook.sheetnames],
    }

    with open(args.output, "w", encoding="utf-8") as fp:
        json.dump(result, fp, ensure_ascii=False, indent=2)
        fp.write("\n")


if __name__ == "__main__":
    main()
